from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from jobseeking_agent.database.models import Job

JOB_EXPORT_HEADERS = [
    "ID",
    "Title",
    "Company",
    "Location",
    "Employment Type",
    "Salary",
    "Status",
    "Platform",
    "Source URL",
    "Official Apply URL",
    "Description",
    "Created At",
]


def export_jobs_to_xlsx(jobs: list[Job], output_path: str | Path) -> Path:
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Jobs"
    sheet.append(JOB_EXPORT_HEADERS)

    for job in jobs:
        sheet.append(
            [
                job.id,
                job.title,
                job.company,
                job.location,
                job.employment_type,
                job.salary,
                job.status,
                job.platform,
                job.job_url,
                job.official_apply_url,
                job.description,
                job.created_at,
            ]
        )

    header_fill = PatternFill("solid", fgColor="EAF2F8")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        column_letter = get_column_letter(column_cells[0].column)
        sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 48)

    sheet.freeze_panes = "A2"
    workbook.save(path)
    return path
